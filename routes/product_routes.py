from datetime import datetime

from flask import Blueprint, jsonify, request
from flask_jwt_extended import get_jwt_identity, jwt_required

import openpyxl
from openpyxl.styles import Font
import os

from app.extensions import db
from models.user_models import RoleEnum, User

from models.product_models import Category, Product

product_blueprint = Blueprint("products", __name__)


@product_blueprint.route("/categories", methods=["POST"])
def create_cat():
    data = request.get_json()

    if not data or not "name" in data:
        return jsonify({"error": "Invalid data"}), 400

    parent_id = data.get("parent_id")

    if parent_id:
        parent_cat = Category.query.get(parent_id)
        if not parent_cat:
            return jsonify({"error": "Parent category not found"}), 404
        else:
            parent_cat = None

    category = Category(name=data["name"], parent_id=parent_id)
    db.session.add(category)
    db.session.commit()

    return (
        jsonify(
            {"Message": "Category created successfully", "category": category.to_dict()}
        ),
        201,
    )


@product_blueprint.route("/categories", methods=["GET"])
def get_cat():
    parent_id = request.args.get("parent_id", type=int)

    if parent_id is not None:
        cats = Category.query.filter_by(parent_id=parent_id).all()
    else:
        cats = Category.query.all()
    return ([cat.to_dict() for cat in cats]), 200


@product_blueprint.route("/categories/<int:cat_id>", methods=["DELETE"])
def delete_cat(cat_id):
    cat = Category.query.get(cat_id)

    if cat is None:
        return jsonify({"error": "Category not found"}), 404

    if cat.children:
        return (
            jsonify(
                {
                    "error": "Category has child categories. Please remove child categories first"
                }
            ),
            400,
        )

    db.session.delete(cat)
    db.session.commit()
    return jsonify({"Message": "Category deleted successfully"}), 200


@product_blueprint.route("/product-create", methods=["POST"])
@jwt_required()
def create_product():
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)

    if not user or user.role.name != RoleEnum.ADMIN:
        return jsonify({"error": "Access denied. Only admins can create products."})
    data = request.get_json()

    if not data or not all(
        pro in data for pro in ["category_id", "name", "title", "price"]
    ):
        return jsonify({"error": "Invalid data"}), 400

    category_id = data.get("category_id")

    if category_id:
        category = Category.query.get(category_id)
        if not category:
            return jsonify({"error": "Category id not found"}), 404
    else:
        category = None

    product = Product(
        name=data["name"],
        title=data["title"],
        price=data["price"],
        category_id=category.id if category else None,
    )
    db.session.add(product)
    db.session.commit()

    write_product_to_xlsx(product)

    return (
        jsonify(
            {"Message": "Product create successfully", "product": product.to_dict()}
        ),
        201,
    )

def write_product_to_xlsx(product):
    file_name = "xlsx_files/product_list.xlsx"

    headers = ["Product ID", "Name", "Title", "Price", "Category ID", "Created_at"]

    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.append(headers)

    worksheet.column_dimensions['A'].width = 8  # Product ID
    worksheet.column_dimensions['B'].width = 25  # Name
    worksheet.column_dimensions['C'].width = 40  # Title
    worksheet.column_dimensions['D'].width = 8  # Price
    worksheet.column_dimensions['E'].width = 8  # Category ID
    worksheet.column_dimensions['F'].width = 20  # Created_at
    font_size = Font(bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = font_size

    if isinstance(product.created_at, str):
        product.created_at = datetime.strptime(product.created_at, "%Y-%m-%d %H:%M")

    worksheet.append([
            product.id,
            product.name,
            product.title,
            product.price,
            product.category_id,
            product.created_at
        ])

    workbook.save(file_name)
    workbook.close()


@product_blueprint.route("/product-list", methods=["GET"])
def product_list():
    products = Product.query.all()
    return ([product.to_dict() for product in products]), 200


@product_blueprint.route("/category/<int:category_id>/products", methods=["GET"])
def get_products_by_category(category_id):
    category = Category.query.get(category_id)

    if not category:
        return jsonify({"error": "Category not found"}), 404

    products = Product.query.filter_by(category_id=category_id).all()

    if not products:
        return jsonify({"error": "No products found in this category"}), 404

    products_list = [product.to_dict() for product in products]

    return jsonify({"category": category.name, "products": products_list}), 200
